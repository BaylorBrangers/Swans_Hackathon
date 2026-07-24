"""Parse medical chronology xlsx files with hyperlink extraction."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

COLUMN_MAP = {
    "Encounter Date": "encounter_date",
    "Primary Provider": "primary_provider",
    "Facility": "facility",
    "Body Parts": "body_parts",
    "Medicine Type": "medicine_type",
    "Record Type": "record_type",
    "Summary": "summary",
    "Link To Pdf": "pdf_url",
}

EXPECTED_COLUMNS = list(COLUMN_MAP.keys())


def _extract_hyperlinks(worksheet) -> list[str | None]:
    """Read hyperlink targets from the Link To Pdf column."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        link_col_idx = list(header_row).index("Link To Pdf") + 1
    except ValueError as exc:
        raise ValueError(
            f"Expected column 'Link To Pdf' not found. Found: {list(header_row)}"
        ) from exc

    urls: list[str | None] = []
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=link_col_idx)
        hyperlink = cell.hyperlink.target if cell.hyperlink else None
        urls.append(hyperlink)

    return urls


def load_chronology_from_bytes(xlsx_bytes: bytes) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load xlsx bytes into a normalized DataFrame with pdf_url hyperlinks."""
    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    worksheet = workbook.active

    pdf_urls = _extract_hyperlinks(worksheet)
    raw_df = pd.read_excel(BytesIO(xlsx_bytes), engine="openpyxl")

    missing = [col for col in EXPECTED_COLUMNS if col not in raw_df.columns]
    if missing:
        raise ValueError(
            f"Missing expected columns: {missing}. Found: {list(raw_df.columns)}"
        )

    df = raw_df[EXPECTED_COLUMNS].copy()
    df = df.rename(columns=COLUMN_MAP)
    df["pdf_url"] = pdf_urls[: len(df)]

    stats: dict[str, Any] = {"total_rows": len(df), "skipped_rows": 0}

    df["encounter_date"] = pd.to_datetime(df["encounter_date"], errors="coerce")
    invalid_mask = df["encounter_date"].isna()
    stats["skipped_rows"] = int(invalid_mask.sum())
    df = df.loc[~invalid_mask].copy()

    for col in ("primary_provider", "facility", "body_parts", "medicine_type", "record_type", "summary"):
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["pdf_url"] = df["pdf_url"].fillna("").astype(str).str.strip()
    df.loc[df["pdf_url"] == "nan", "pdf_url"] = ""

    df = df.sort_values("encounter_date", ascending=False).reset_index(drop=True)
    return df, stats


def load_chronology_from_path(path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load a local xlsx file into a normalized DataFrame."""
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Local xlsx not found: {file_path}")
    return load_chronology_from_bytes(file_path.read_bytes())


def split_multi_value(value: str, separator: str) -> list[str]:
    """Split semicolon- or comma-separated filter values into trimmed parts."""
    if not value or value.lower() == "nan":
        return []
    return [part.strip() for part in value.split(separator) if part.strip()]


def unique_body_parts(series: pd.Series) -> list[str]:
    """Return sorted unique body parts from comma-separated values."""
    parts: set[str] = set()
    for value in series:
        parts.update(split_multi_value(str(value), ","))
    return sorted(parts)


def unique_providers(series: pd.Series) -> list[str]:
    """Return sorted unique provider names from semicolon-separated values."""
    providers: set[str] = set()
    for value in series:
        providers.update(split_multi_value(str(value), ";"))
    return sorted(providers)
